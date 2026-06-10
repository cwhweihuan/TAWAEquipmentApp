#!/usr/bin/env python3
"""Extract the WeihuanTAWA master equipment list + sample stores from the source xlsx.

Produces:
  data/seed/equipment.json   - 210 master equipment records (with PDF refs)
  data/seed/stores.json      - the 3 existing store configurations
  data/seed/pdfs.json        - unique Drive PDF links to download
"""
import json, os, re
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, "source", "equipment.xlsx")
OUT = os.path.join(HERE, "seed")
os.makedirs(OUT, exist_ok=True)

DEPARTMENTS = ["MEAT", "SEAFOOD", "PRODUCE", "GROCERY", "BAKERY",
               "HOT DELI", "WAREHOUSE", "GENERAL MARKET", "SUBTENANTS"]

# 1-indexed column map for the MASTER sheet
COL = {
    "master": 1,
    # dept flag cols 3..11, dept item# cols 12..20
    "description": 21,
    "specLink": 22,
    "manufacturer": 23,
    "model": 24,
    "supplyBy": 25,
    "installBy": 26,
    "power": 27, "height": 28, "nema": 29,
    "dataPhone": 30,
    "waterCold": 31, "waterHot": 32, "waterElev": 33,
    "gasPipe": 34, "gasBtu": 35, "gasElev": 36,
    "floorSink": 37, "remarks": 38,
}

def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def drive_id(url):
    m = re.search(r"/d/([A-Za-z0-9_-]+)", url or "")
    return m.group(1) if m else None

def extract_master():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    wb_f = openpyxl.load_workbook(SRC, data_only=False)  # for hyperlinks
    ws = wb[wb.sheetnames[0]]
    ws_f = wb_f[wb_f.sheetnames[0]]

    equipment, pdfs = [], {}
    master_counter = 0
    for r in range(4, ws.max_row + 1):
        desc = clean(ws.cell(row=r, column=COL["description"]).value)
        if not desc:
            continue
        # skip the legend rows at the bottom (e.g. "G.C.: GENERAL CONTRACTOR")
        if re.match(r"^[A-Z]\.[A-Z]\.:", desc):
            continue
        master_counter += 1  # master# is formula-driven; use stable running index

        # department flags + per-dept item numbers
        depts, dept_items = [], {}
        for i, dept in enumerate(DEPARTMENTS):
            flag = clean(ws.cell(row=r, column=3 + i).value)
            item_no = clean(ws.cell(row=r, column=12 + i).value)
            if flag and flag.lower() == "x":
                depts.append(dept)
            if item_no:
                dept_items[dept] = item_no.replace("\n", " / ")

        # PDF hyperlink lives on the spec-link cell in the formula workbook
        pdf_id = None
        cell = ws_f.cell(row=r, column=COL["specLink"])
        if cell.hyperlink and cell.hyperlink.target and "drive.google" in str(cell.hyperlink.target):
            url = cell.hyperlink.target
            did = drive_id(url)
            if did:
                pdf_id = did
                pdfs[did] = {"driveId": did, "url": url, "filename": f"{did}.pdf"}

        equipment.append({
            "masterItemNo": master_counter,
            "description": desc,
            "departments": depts,
            "departmentItems": dept_items,
            "manufacturer": clean(ws.cell(row=r, column=COL["manufacturer"]).value),
            "model": clean(ws.cell(row=r, column=COL["model"]).value),
            "supplyBy": clean(ws.cell(row=r, column=COL["supplyBy"]).value),
            "installBy": clean(ws.cell(row=r, column=COL["installBy"]).value),
            "power": clean(ws.cell(row=r, column=COL["power"]).value),
            "height": clean(ws.cell(row=r, column=COL["height"]).value),
            "nema": clean(ws.cell(row=r, column=COL["nema"]).value),
            "dataPhone": clean(ws.cell(row=r, column=COL["dataPhone"]).value),
            "waterCold": clean(ws.cell(row=r, column=COL["waterCold"]).value),
            "waterHot": clean(ws.cell(row=r, column=COL["waterHot"]).value),
            "waterElev": clean(ws.cell(row=r, column=COL["waterElev"]).value),
            "gasPipe": clean(ws.cell(row=r, column=COL["gasPipe"]).value),
            "gasBtu": clean(ws.cell(row=r, column=COL["gasBtu"]).value),
            "gasElev": clean(ws.cell(row=r, column=COL["gasElev"]).value),
            "floorSink": clean(ws.cell(row=r, column=COL["floorSink"]).value),
            "remarks": clean(ws.cell(row=r, column=COL["remarks"]).value),
            "pdfDriveId": pdf_id,
        })
    return equipment, list(pdfs.values())

def extract_stores():
    """Each non-master sheet is an existing store configuration."""
    wb = openpyxl.load_workbook(SRC, data_only=True)
    wb_f = openpyxl.load_workbook(SRC, data_only=False)
    stores = []
    for name in wb.sheetnames[1:]:
        ws, ws_f = wb[name], wb_f[name]
        # store sheet columns (1-indexed): A=dept item#, B=room, C=propose-new,
        # D=spec.link, E=quantity, F=schedule#, G=description, H=mfr, I=model ...
        items = []
        for r in range(3, ws.max_row + 1):
            desc = clean(ws.cell(row=r, column=7).value)
            if not desc:
                continue
            pdf_id = None
            cell = ws_f.cell(row=r, column=4)
            if cell.hyperlink and cell.hyperlink.target and "drive.google" in str(cell.hyperlink.target):
                pdf_id = drive_id(cell.hyperlink.target)
            items.append({
                "deptItemNo": clean(ws.cell(row=r, column=1).value),
                "room": clean(ws.cell(row=r, column=2).value),
                "proposeNew": clean(ws.cell(row=r, column=3).value),
                "quantity": clean(ws.cell(row=r, column=5).value),
                "scheduleNo": clean(ws.cell(row=r, column=6).value),
                "description": desc,
                "manufacturer": clean(ws.cell(row=r, column=8).value),
                "model": clean(ws.cell(row=r, column=9).value),
                "supplyBy": clean(ws.cell(row=r, column=10).value),
                "installBy": clean(ws.cell(row=r, column=11).value),
                "pdfDriveId": pdf_id,
            })
        # parse "#95 PORTLAND, OR" -> number + location
        m = re.match(r"#?(\S+)\s+(.*)", name)
        stores.append({
            "name": name,
            "number": m.group(1) if m else name,
            "location": m.group(2) if m else "",
            "items": items,
        })
    return stores

if __name__ == "__main__":
    equipment, pdfs = extract_master()
    stores = extract_stores()
    json.dump(equipment, open(os.path.join(OUT, "equipment.json"), "w"), indent=2, ensure_ascii=False)
    json.dump(pdfs, open(os.path.join(OUT, "pdfs.json"), "w"), indent=2, ensure_ascii=False)
    json.dump(stores, open(os.path.join(OUT, "stores.json"), "w"), indent=2, ensure_ascii=False)
    print(f"equipment: {len(equipment)} records")
    print(f"  with PDFs: {sum(1 for e in equipment if e['pdfDriveId'])}")
    print(f"  with departments: {sum(1 for e in equipment if e['departments'])}")
    print(f"unique PDFs: {len(pdfs)}")
    print(f"stores: {len(stores)} -> " + ", ".join(f"{s['name']}({len(s['items'])})" for s in stores))
